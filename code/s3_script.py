import pandas as pd
import io
import boto3
import openpyxl
import time
from openpyxl.styles import Border, Side
from tempfile import NamedTemporaryFile

#os.chdir("/users/morgana.sartor/desktop/development/github/claro_free_report_poc")

class ClaroFreeAutomateReports:

    def __init__(self):
        self.source_database_name = 'claro_free'
        self.source_table_name = 'rel_campaign'
        self.s3_bucket = 'prezao-free-report-test'
        self.s3_output_querys = f's3://{self.s3_bucket}/source-querys'
        self.s3_target_url = f's3://{self.s3_bucket}/reports'
        self.s3_model_url = f's3://{self.s3_bucket}/model'
        self.region_name = 'us-east-1'

        self.athena_client = boto3.client('athena', region_name=self.region_name)
        self.s3_resource = boto3.resource('s3', region_name=self.region_name)
        self.s3_client = boto3.client('s3', region_name=self.region_name)

        # Valores de células do excel
        self.report_infos_overview_column = 3
        self.report_infos_start_line = 6
        
        self.report_infos_consolidaded_column = 13
        self.report_infos_consolidaded_start_line = 6
        
        self.data_table_start_column = 2
        self.data_table_last_column = 7
        self.data_table_start_line = 6
        self.data_table_total_line = 5

        self.limit_index = 28
        
        self.thin = Side(border_style="thin", color="000000")

    def get_filename_from_athena_query(self, query):
        filename = ''
        try:
            response = self.athena_client.start_query_execution(
                QueryString = query,
                    QueryExecutionContext={
                    'Database': self.source_database_name
                    },
                    ResultConfiguration={
                    'OutputLocation': self.s3_output_querys,
                    }
            )
            filename = response['QueryExecutionId']
            print('Execution ID: ' + response['QueryExecutionId'])

            query_status = None
            while query_status == 'QUEUED' or query_status == 'RUNNING' or query_status is None:
                query_status = self.athena_client.get_query_execution(QueryExecutionId=response["QueryExecutionId"])['QueryExecution']['Status']['State']
                print(f'Query status: {query_status}')
                print('...')
                if query_status == 'FAILED' or query_status == 'CANCELLED':
                    raise Exception('Athena query with the string "{}" failed or was cancelled'.format(self.query))
                time.sleep(10)
            print(f'Query {query} finished.')
        except Exception as e:
            print(e)

        return filename

    def get_df_from_athena_query(self, query):
        filename = self.get_filename_from_athena_query(query)
        result = ''

        response = self.s3_resource \
                    .Bucket(self.s3_bucket) \
                    .Object(key='source-querys/' + filename + '.csv') \
                    .get()

        result = pd.read_csv(io.BytesIO(response['Body'].read()), encoding='utf8') 

        return result

    def all_reports_on_table(self):
        query = 'select distinct name as report from claro_free.rel_campaign'
        self.all_reports_on_table = self.get_df_from_athena_query(query)

    def all_reports_on_s3(self):
        s3_bucket = self.s3_resource.Bucket(self.s3_bucket)
        folder = 'reports/'
        self.reports_already_created = [
            f.key.split(folder + "/")[1] 
            for f in s3_bucket.objects.filter(Prefix=folder).all()
            if len(f.key.split(folder + "/")) > 1
        ]

    def get_reports_to_be_created(self):
        reports_to_create = list(set(self.all_reports_on_table['report']) - set(self.reports_already_created))
        return reports_to_create

    def get_reports_to_be_increased(self):
        reports_to_increase = list(set(self.all_reports_on_table['report']).intersection(self.reports_already_created))
        return reports_to_increase

    def create_report(self, campaign_name, wb):
        print(f'Report: {campaign_name}')
        
        query = f"select * from claro_free.rel_campaign where name = '{campaign_name}'"
        results = self.get_df_from_athena_query(query)
        
        print(f'Data of {campaign_name}: \n{results}')
        
        results.sort_values(by=['date'], ascending=True, inplace=True)

        # Inserir dados na primeira planilha
        sheet_ranges = wb['Dashboard']

        #advertiser = report_data.advertiser.unique()
        #sheet_ranges.cell(row=report_infos_start_line, column=report_infos_overview_column).value = advertiser[0]    
        campaign = campaign_name
        sheet_ranges.cell(row=self.report_infos_start_line+1, column=self.report_infos_overview_column).value = campaign[0]
        volume = results.volume.unique()
        sheet_ranges.cell(row=self.report_infos_start_line+2, column=self.report_infos_overview_column).value = volume[0] 
        value = results.cpm.unique()
        sheet_ranges.cell(row=self.report_infos_start_line+3, column=self.report_infos_overview_column).value = value[0]
        date_data = results['date'].reset_index(drop=True)
        sheet_ranges.cell(row=self.report_infos_start_line+6, column=self.report_infos_overview_column).value = date_data[0]
    
        size_data = len(date_data)
        finish_date = date_data[size_data - 1]
    
        impression = results['impression'].reset_index(drop=True)
        clicked = results['clicked'].reset_index(drop=True)
        complete = results['complete'].reset_index(drop=True)

        # Inserir dados na segunda planilha
        sheet_ranges = wb['Dados']

        last_index = self.data_table_start_line + len(date_data)

        # Limpa todas as células do modelo
        for cell in range(self.data_table_start_line, self.limit_index):
            sheet_ranges.cell(row=cell, column=2).value = ''

        # Retira formato da última célula
        for col in range(self.data_table_start_column, self.data_table_last_column):
            sheet_ranges.cell(row=self.limit_index, column=col).border = None

        # Preenche dados
        for cell in range(self.data_table_start_line, last_index):
        
            sheet_ranges.cell(row=cell, column=self.data_table_start_column).value = date_data[cell - self.data_table_start_line]
            sheet_ranges.cell(row=cell, column=self.data_table_start_column+1).value = impression[cell - self.data_table_start_line]
            sheet_ranges.cell(row=cell, column=self.data_table_start_column+2).value = clicked[cell - self.data_table_start_line]
            sheet_ranges.cell(row=cell, column=self.data_table_start_column+3).value = f'=D{cell}/C{cell}'
            sheet_ranges.cell(row=cell, column=self.data_table_start_column+4).value = complete[cell - self.data_table_start_line]
            sheet_ranges.cell(row=cell, column=self.data_table_start_column+5).value = f'=F{cell}/D{cell}'
    
            # Mantem a formatação
            sheet_ranges.cell(row=cell, column=self.data_table_start_column).border = Border(left=self.thin)
            sheet_ranges.cell(row=cell, column=self.data_table_last_column).border = Border(right=self.thin)
        
        sheet_ranges.cell(row=self.data_table_total_line, column=self.data_table_start_column+1).value = f'=SUM(C{self.data_table_start_line}:C{last_index - 1})'
        sheet_ranges.cell(row=self.data_table_total_line, column=self.data_table_start_column+2).value = f'=SUM(D{self.data_table_start_line}:D{last_index - 1})'
        sheet_ranges.cell(row=self.data_table_total_line, column=self.data_table_start_column+3).value = f'=D{last_index - 1}/C{last_index - 1}'
        sheet_ranges.cell(row=self.data_table_total_line, column=self.data_table_start_column+4).value = f'=SUM(F{self.data_table_start_line}:F{last_index - 1})'
        sheet_ranges.cell(row=self.data_table_total_line, column=self.data_table_start_column+5).value = f'=F{last_index - 1}/D{last_index - 1}'

        # Adiciona formatação na última linha
        for col in range(self.data_table_start_column, self.data_table_last_column+1):
            if col == self.data_table_start_column:
                sheet_ranges.cell(row=cell, column=col).border = Border(left=self.thin, bottom=self.thin)
            elif col == self.data_table_last_column:
                sheet_ranges.cell(row=cell, column=col).border = Border(right=self.thin, bottom=self.thin)
            else:
                sheet_ranges.cell(row=cell, column=col).border = Border(bottom=self.thin)
    
        sheet_ranges = wb['Dashboard']
        sheet_ranges.cell(row=self.report_infos_consolidaded_start_line, column=self.report_infos_consolidaded_column).value = f'=Dados!C{self.data_table_total_line}'
        sheet_ranges.cell(row=self.report_infos_consolidaded_start_line+1, column=self.report_infos_consolidaded_column).value = f'=Dados!D{self.data_table_total_line}'
        sheet_ranges.cell(row=self.report_infos_consolidaded_start_line+2, column=self.report_infos_consolidaded_column).value = f'=Dados!F{self.data_table_total_line}'
        sheet_ranges.cell(row=self.report_infos_consolidaded_start_line+3, column=self.report_infos_consolidaded_column).value = f'=Dados!G{self.data_table_total_line}'

        folder_name = campaign_name.replace('/', '.').replace(' ','-')
        file_name = f'reports/{folder_name}/{folder_name}({finish_date}).xlsx'
        self.s3_client.put_object(Bucket=self.s3_bucket, Key=file_name)
        with NamedTemporaryFile() as tmp:
            temp_file = f'tmp/tmp.xlsx'
            wb.save(temp_file)
            self.s3_resource.Bucket(self.s3_bucket).upload_file(Filename=temp_file, Key=file_name)

        print(f'Created report {report} in s3://{self.s3_bucket}/{file_name}\n')


if __name__ == "__main__":
    claro_free_automate_reports = ClaroFreeAutomateReports()
    claro_free_automate_reports.all_reports_on_table()
    claro_free_automate_reports.all_reports_on_s3()

    print(f'Reports on table: {claro_free_automate_reports.all_reports_on_table}\n')
    print(f'Reports already created: {claro_free_automate_reports.reports_already_created}\n')

    reports_to_be_created = claro_free_automate_reports.get_reports_to_be_created()
    reports_to_be_increased = claro_free_automate_reports.get_reports_to_be_increased()
        
    print(f'Reports to create: {reports_to_be_created}\n')
    print(f'Reports to increase: {reports_to_be_increased}\n')

    model_object = claro_free_automate_reports.s3_resource.Bucket(claro_free_automate_reports.s3_bucket).Object(key='model').get()
    model_path = io.BytesIO(model_object['Body'].read())

    for report in reports_to_be_created:
        workbook = openpyxl.load_workbook(model_path)
        claro_free_automate_reports.create_report(report, workbook)
        workbook.close()