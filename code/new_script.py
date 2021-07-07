import pandas as pd
import os
import glob
import openpyxl
from openpyxl.styles import Border, Side

os.chdir("/users/morgana.sartor/desktop/development/github/claro_free_report_poc")

# Importa nova fonte de dados com dados mais recentes
df = pd.read_csv('data/new_data_source.csv')

# Limpa e formata os dados
df[['advertiser', 'report']] = ['', '']
new = df.name.str.split(pat='|',n=2,expand=True)

try:
    df['name'] = new[0]
    df['advertiser'] = new[1]
    df['report'] = new[2]
except:
    df['name'] = df['name']
    
df['name'] = df['name'].str.strip()
df['advertiser'] = df['advertiser'].str.strip()
df['report'] = df['report'].str.strip()

report_df = df[df['report']=='X']
report_df['name'] = report_df['name'].str.replace('/', '-')

# Verifica quais já existem
existents_reports = os.listdir("/users/morgana.sartor/desktop/development/github/claro_free_report_poc/report")

# Lista dos que necessitam ser criados
need_reports = report_df.name.unique()

reports_to_create = list(set(need_reports) - set(existents_reports))
reports_to_increase = list(set(need_reports).intersection(existents_reports))

print(f'Reports to create: {reports_to_create}\n')
print(f'Reports to increase: {reports_to_increase}\n')

# Valores de células do excel
report_infos_overview_column = 3
report_infos_start_line = 6

report_infos_consolidaded_column = 13
report_infos_consolidaded_start_line = 6

data_table_start_column = 2
data_table_last_column = 7
data_table_start_line = 6
data_table_total_line = 5

limit_index = 28

# Auxiliares para formatação
thin = Side(border_style="thin", color="000000")

for report in reports_to_create:
    report_data = report_df[report_df['name']==report]
    report_data.sort_values(by=['date'], ascending=True, inplace=True)

    wb = openpyxl.load_workbook('reportmodel/model.xlsx')

    # Inserir dados na primeira planilha
    sheet_ranges = wb['Dashboard']

    advertiser = report_data.advertiser.unique()
    sheet_ranges.cell(row=report_infos_start_line, column=report_infos_overview_column).value = advertiser[0]    
    campaign = report_data.name.unique()
    sheet_ranges.cell(row=report_infos_start_line+1, column=report_infos_overview_column).value = campaign[0]
    volume = report_data.volume.unique()
    sheet_ranges.cell(row=report_infos_start_line+2, column=report_infos_overview_column).value = volume[0] 
    value = report_data.cpm.unique()
    sheet_ranges.cell(row=report_infos_start_line+3, column=report_infos_overview_column).value = value[0]
    date_data = report_data['date'].reset_index(drop=True)
    sheet_ranges.cell(row=report_infos_start_line+6, column=report_infos_overview_column).value = date_data[0]
    
    size_data = len(date_data)
    finish_date = date_data[size_data - 1]
    
    impression = report_data['impression'].reset_index(drop=True)
    clicked = report_data['clicked'].reset_index(drop=True)
    complete = report_data['complete'].reset_index(drop=True)

    # Inserir dados na segunda planilha
    sheet_ranges = wb['Dados']
    
    last_index = data_table_start_line + len(date_data)
    
    # Limpa todas as células do modelo
    for cell in range(data_table_start_line, limit_index):
        sheet_ranges.cell(row=cell, column=2).value = ''

    # Retira formato da última célula
    for col in range(data_table_start_column, data_table_last_column):
    	sheet_ranges.cell(row=limit_index, column=col).border = None

    # Preenche dados
    for cell in range(data_table_start_line, last_index):
        
        sheet_ranges.cell(row=cell, column=data_table_start_column).value = date_data[cell - data_table_start_line]
        sheet_ranges.cell(row=cell, column=data_table_start_column+1).value = impression[cell - data_table_start_line]
        sheet_ranges.cell(row=cell, column=data_table_start_column+2).value = clicked[cell - data_table_start_line]
        sheet_ranges.cell(row=cell, column=data_table_start_column+3).value = f'=D{cell}/C{cell}'
        sheet_ranges.cell(row=cell, column=data_table_start_column+4).value = complete[cell - data_table_start_line]
        sheet_ranges.cell(row=cell, column=data_table_start_column+5).value = f'=F{cell}/D{cell}'

        # Mantem a formatação
        sheet_ranges.cell(row=cell, column=data_table_start_column).border = Border(left=thin)
        sheet_ranges.cell(row=cell, column=data_table_last_column).border = Border(right=thin)
        
    sheet_ranges.cell(row=data_table_total_line, column=data_table_start_column+1).value = f'=SUM(C{data_table_start_line}:C{last_index - 1})'
    sheet_ranges.cell(row=data_table_total_line, column=data_table_start_column+2).value = f'=SUM(D{data_table_start_line}:D{last_index - 1})'
    sheet_ranges.cell(row=data_table_total_line, column=data_table_start_column+3).value = f'=D{last_index - 1}/C{last_index - 1}'
    sheet_ranges.cell(row=data_table_total_line, column=data_table_start_column+4).value = f'=SUM(F{data_table_start_line}:F{last_index - 1})'
    sheet_ranges.cell(row=data_table_total_line, column=data_table_start_column+5).value = f'=F{last_index - 1}/D{last_index - 1}'

    # Adiciona formatação na última linha
    for col in range(data_table_start_column, data_table_last_column+1):
    	if col == data_table_start_column:
    		sheet_ranges.cell(row=cell, column=col).border = Border(left=thin, bottom=thin)
    	elif col == data_table_last_column:
    		sheet_ranges.cell(row=cell, column=col).border = Border(right=thin, bottom=thin)
    	else:
    		sheet_ranges.cell(row=cell, column=col).border = Border(bottom=thin)
    
    sheet_ranges = wb['Dashboard']
    sheet_ranges.cell(row=report_infos_consolidaded_start_line, column=report_infos_consolidaded_column).value = f'=Dados!C{data_table_total_line}'
    sheet_ranges.cell(row=report_infos_consolidaded_start_line+1, column=report_infos_consolidaded_column).value = f'=Dados!D{data_table_total_line}'
    sheet_franges.cell(row=report_infos_consolidaded_start_line+2, column=report_infos_consolidaded_column).value = f'=Dados!F{data_table_total_line}'
    sheet_ranges.cell(row=report_infos_consolidaded_start_line+3, column=report_infos_consolidaded_column).value = f'=Dados!G{data_table_total_line}'

    os.mkdir('report/' + report)
    file_name = f'report/{report}/{report}({finish_date}).xlsx'
    wb.save(file_name)

    print(f'Created report {report} in {file_name}\n')

for report in reports_to_increase:
    report_data = report_df[report_df['name']==report]
    report_data.sort_values(by=['date'], ascending=True, inplace=True)
    
    date_data = report_data['date'].reset_index(drop=True)
    size_data = len(date_data)
    finish_date = date_data[size_data - 1]
    
    file_path = f'report/{report}'
    files = glob.glob(file_path + '/*.xlsx')
    max_file = max(files, key=os.path.getctime)
    wb = openpyxl.load_workbook(max_file)
    
    # Inserir dados na segunda planilha
    sheet_ranges = wb['Dados']
    
    impression = report_data['impression'].reset_index(drop=True)
    clicked = report_data['clicked'].reset_index(drop=True)
    complete = report_data['complete'].reset_index(drop=True)
    
    last_line_with_data = sheet_ranges.max_row
    for cell in range(data_table_start_line, sheet_ranges.max_row):
    	if sheet_ranges.cell(row=cell, column=data_table_start_column).value == None:
    		last_line_with_data = cell - 1
    		break;

    print(f'Last line with data: {last_line_with_data}\n')

    last_index = last_line_with_data + len(date_data)
    
    # Retira formato da última célula
    for col in range(data_table_start_column, data_table_last_column):
    	sheet_ranges.cell(row=last_line_with_data, column=col).border = Border(bottom=None)

    # Adiciona dados novos
    for cell in range(last_line_with_data, last_index):
        sheet_ranges.cell(row=cell, column=data_table_start_column).value = date_data[cell - last_line_with_data]
        sheet_ranges.cell(row=cell, column=data_table_start_column+1).value = impression[cell - last_line_with_data]
        sheet_ranges.cell(row=cell, column=data_table_start_column+2).value = clicked[cell - last_line_with_data]
        sheet_ranges.cell(row=cell, column=data_table_start_column+3).value = f'=D{cell}/C{cell}'
        sheet_ranges.cell(row=cell, column=data_table_start_column+4).value = complete[cell - last_line_with_data]
        sheet_ranges.cell(row=cell, column=data_table_start_column+5).value = f'=F{cell}/D{cell}'

        # Mantem a formatação
        sheet_ranges.cell(row=cell, column=data_table_start_column).border = Border(left=thin)
        sheet_ranges.cell(row=cell, column=data_table_last_column).border = Border(right=thin)

    sheet_ranges.cell(row=data_table_total_line, column=data_table_start_column+1).value = f'=SUM(C{data_table_start_line}:C{last_index - 1})'
    sheet_ranges.cell(row=data_table_total_line, column=data_table_start_column+2).value = f'=SUM(D{data_table_start_line}:D{last_index - 1})'
    sheet_ranges.cell(row=data_table_total_line, column=data_table_start_column+3).value = f'=D{last_index - 1}/C{last_index - 1}'
    sheet_ranges.cell(row=data_table_total_line, column=data_table_start_column+4).value = f'=SUM(F{data_table_start_line}:F{last_index - 1})'
    sheet_ranges.cell(row=data_table_total_line, column=data_table_start_column+5).value = f'=F{last_index - 1}/D{last_index - 1}'
    
    # Adiciona formatação na última linha
    for col in range(data_table_start_column, data_table_last_column+1):
    	if col == data_table_start_column:
    		sheet_ranges.cell(row=cell, column=col).border = Border(left=thin, bottom=thin)
    	elif col == data_table_last_column:
    		sheet_ranges.cell(row=cell, column=col).border = Border(right=thin, bottom=thin)
    	else:
    		sheet_ranges.cell(row=cell, column=col).border = Border(bottom=thin)

    sheet_ranges = wb['Dashboard']
    sheet_ranges['M6'] = f'=Dados!C{data_table_total_line}'
    sheet_ranges['M7'] = f'=Dados!D{data_table_total_line}'
    sheet_ranges['M8'] = f'=Dados!F{data_table_total_line}'
    sheet_ranges['M9'] = f'=Dados!G{data_table_total_line}'
    
    file_name = f'report/{report}/{report}({finish_date}).xlsx'
    wb.save(file_name)

    print(f'Increased report {report} in {file_name}\n')